import openpyxl

workbook_path = "C:\\Users\\rames\\Downloads\\excel_workbook.xlsx"
book = openpyxl.load_workbook(workbook_path)
sheet = book.active

my_dict = {"testcase1": {"Firstname": "Jeeva", "lastname": "Nandham", "email": "jeeva474@gmail.com"},
           "testcase2": {"Firstname": "Reshma", "lastname": "Kothandaraman", "email": "reshmak3333@gmail.com"},
           "testcase3": {"Firstname": "Ramesh", "lastname": "babu", "email": "mesh7271@gmail.com"},
           "testcase4": {"Firstname": "krithik", "lastname": "jeeva", "email": "krithik474@gmail.com"},
           "testcase5": {"Firstname": "gowtham", "lastname": "babu", "email": "gowtham_babu@gmail.com"},
           }
print(sheet.max_row)
print(sheet.max_column)

for row in range(2, sheet.max_row + 1):
    row_value = sheet.cell(row=row, column=1).value
    if row_value in my_dict:
        for column in range(2, sheet.max_column+1):
            column_header = sheet.cell(row=1, column=column).value
            if column_header in my_dict[row_value]:
                dict_value = my_dict[row_value][column_header]
                sheet.cell(row=row, column=column).value = dict_value

for row in range(2, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        print(sheet.cell(row=row, column=column).value)

book.save(workbook_path)
