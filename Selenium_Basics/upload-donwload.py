from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import openpyxl

filepath = "C:\\Users\\rames\\Downloads\\download.xlsx"

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--ignore-certificate-erros")

driver = webdriver.Chrome(options=chrome_options)
driver.implicitly_wait(5)

driver.get("https://rahulshettyacademy.com/upload-download-test/")

driver.find_element(By.ID, "downloadButton").click()
assert os.path.exists(filepath), "File not found in download directory"

def update_excel_data(filepath):
    book = openpyxl.load_workbook(filepath)
    sheet = book.active

    for row in range(2,sheet.max_row):
        if sheet.cell(row=row, column=2).value == "Apple":
            sheet.cell(row=row, column=4).value = 400
    book.save(filepath)

update_excel_data(filepath)

driver.find_element(By.CSS_SELECTOR, "input[type='file']").send_keys(filepath)
wait = WebDriverWait(driver, 5)
toast_locator = (By.XPATH, "/html[1]/body[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[2]")
wait.until(EC.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

apple_price = driver.find_element(By.XPATH, "//div[text()='Apple']/parent::div/parent::div/div[@id='cell-4-undefined']").text
papaya_price = driver.find_element(By.XPATH,"//div[text()='Papaya']/parent::div/parent::div/div[@id='cell-4-undefined']").text
orange_seasons = driver.find_elements(By.XPATH, "//div[text()='Orange']/parent::div/parent::div/div[@id='cell-5-undefined']")
orange_fruit_seasons = []
for season in orange_seasons:
    orange_fruit_seasons.append(season.text)
print(apple_price)
print(papaya_price)
print(orange_fruit_seasons)